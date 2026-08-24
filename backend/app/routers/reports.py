import csv
import io
from datetime import date, datetime
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from .. import models
from ..database import get_db
from ..deps import require_employee
from .dashboard import _sum_by_resource

router = APIRouter()

# Mirrors the frontend's statusInfo() precedence (is_approved -> latest
# revision/rejected -> current approval_stage) so report filtering matches
# exactly what the status badge shows on screen.
STATUS_LABELS = {
    "approved": "Согласовано",
    "revision": "На пересмотре",
    "rejected": "Отклонено",
    "pending_head": "Ожидает согласования руководителя",
    "pending_teamlead": "Ожидает согласования TeamLead",
    "pending_top": "Ожидает согласования топ-менеджмента",
}

REPORT_HEADERS = [
    "Название", "Описание", "Подразделение", "Чемпион", "Статус",
    "Дата начала", "Дата окончания",
    "Человеческие ресурсы (план/факт)", "Технические ресурсы",
    "Выгода в месяц, ₽", "Окупаемость, мес.",
]


def _status_code(ini: models.Initiative) -> str:
    if ini.is_approved:
        return "approved"
    latest = max(ini.approvals, key=lambda a: a.created_at) if ini.approvals else None
    if latest and latest.status == "revision":
        return "revision"
    if latest and latest.status == "rejected":
        return "rejected"
    if ini.approval_stage == "teamlead":
        return "pending_teamlead"
    if ini.approval_stage == "top":
        return "pending_top"
    return "pending_head"


def _base_query(db: Session):
    return db.query(models.Initiative).options(
        joinedload(models.Initiative.champion),
        joinedload(models.Initiative.department),
        joinedload(models.Initiative.resource_entries).joinedload(models.ResourceEntry.resource),
        joinedload(models.Initiative.benefits).joinedload(models.Benefit.resource),
        joinedload(models.Initiative.cost_logs).joinedload(models.CostLog.resource),
        joinedload(models.Initiative.approvals),
    )


def _in_range(value: Optional[date], date_from: Optional[date], date_to: Optional[date]) -> bool:
    if value is None:
        return False
    if date_from and value < date_from:
        return False
    if date_to and value > date_to:
        return False
    return True


def _fmt_qty(q: float) -> str:
    return f"{q:g}"


def _human_resource_summary(ini: models.Initiative) -> str:
    fact_by_resource: dict = {}
    for log in ini.cost_logs:
        fact_by_resource[log.resource_id] = fact_by_resource.get(log.resource_id, 0) + log.quantity
    parts = []
    for e in ini.resource_entries:
        if not e.is_planned or not e.resource or e.resource.category != "human":
            continue
        fact = fact_by_resource.get(e.resource_id, 0)
        parts.append(f"{e.resource.name}: план {_fmt_qty(e.quantity)} / факт {_fmt_qty(fact)} {e.resource.unit}".strip())
    return "; ".join(parts)


def _tech_resource_summary(ini: models.Initiative) -> str:
    parts = [
        f"{e.resource.name}: {_fmt_qty(e.quantity)} {e.resource.unit}".strip()
        for e in ini.resource_entries
        if e.is_planned and e.resource and e.resource.category == "tech"
    ]
    return "; ".join(parts)


def _planned_cost_money(ini: models.Initiative) -> float:
    return sum(e.quantity * (e.resource.rate if e.resource else 0) for e in ini.resource_entries if e.is_planned)


def _benefit_money(ini: models.Initiative) -> float:
    return sum(b.quantity * (b.resource.rate if b.resource else 0) for b in ini.benefits)


def _payback_months(ini: models.Initiative) -> Optional[float]:
    benefit = _benefit_money(ini)
    if benefit <= 0:
        return None
    return _planned_cost_money(ini) / benefit


def _row_values(ini: models.Initiative) -> list:
    payback = _payback_months(ini)
    return [
        ini.title,
        ini.description or "",
        ini.department.name if ini.department else "",
        ini.champion.full_name if ini.champion else "",
        STATUS_LABELS[_status_code(ini)],
        ini.start_date.isoformat() if ini.start_date else "",
        ini.end_date.isoformat() if ini.end_date else "",
        _human_resource_summary(ini),
        _tech_resource_summary(ini),
        round(_benefit_money(ini), 2),
        round(payback, 2) if payback is not None else "",
    ]


def _summary_rows(initiatives: list) -> list:
    """Portfolio-level totals for exactly the initiatives in this report —
    one row per line, single non-empty cell each (matches the per-initiative
    table's column count so both xlsx and csv writers can just append them)."""
    resources_planned = _sum_by_resource(
        (e.resource, e.quantity) for ini in initiatives for e in ini.resource_entries if e.is_planned
    )
    total_benefit = sum(_benefit_money(ini) for ini in initiatives)
    paybacks = [p for p in (_payback_months(ini) for ini in initiatives) if p is not None]
    # Portfolio payback = sum of each initiative's own payback period — same
    # convention as the "Портфель" dashboard's total_payback_months.
    total_payback = sum(paybacks)

    rows = [
        [],
        ["Итого по отчёту"],
        [f"Инициатив в отчёте: {len(initiatives)}"],
        ["Сумма запрашиваемых ресурсов:"],
    ]
    if resources_planned:
        for r in resources_planned:
            rows.append([f"  {r['resource']}: {r['total']:g} {r['unit']}"])
    else:
        rows.append(["  Нет данных"])
    rows.append([f"Суммарная выгода в месяц, ₽: {round(total_benefit, 2)}"])
    rows.append([f"Суммарная окупаемость, мес.: {round(total_payback, 2)}"])
    return rows


def _filter_initiatives(
    db: Session,
    employee: models.Employee,
    date_from: Optional[date],
    date_to: Optional[date],
    status: Optional[str],
    department_id: Optional[int],
) -> list:
    if status is not None and status not in STATUS_LABELS:
        raise HTTPException(status_code=400, detail=f"Статус должен быть одним из: {', '.join(STATUS_LABELS)}")

    q = _base_query(db)

    # Scoping mirrors the rest of the app: champions only ever see their own
    # initiatives, heads only their own department's. The department filter
    # is only meaningful (and only applied) for PM/top, who see everything.
    if employee.role == "champion":
        q = q.filter(models.Initiative.champion_id == employee.id)
    elif employee.role == "head":
        q = q.filter(models.Initiative.department_id == employee.department_id)
    elif employee.role in ("pm", "top"):
        if department_id:
            q = q.filter(models.Initiative.department_id == department_id)
    else:
        raise HTTPException(status_code=403, detail="Раздел «Отчёты» недоступен для этой роли")

    initiatives = q.order_by(models.Initiative.created_at).all()

    # An initiative matches the date range if its start OR end date falls
    # inside it — each bound is optional and independent.
    if date_from or date_to:
        initiatives = [
            ini for ini in initiatives
            if _in_range(ini.start_date, date_from, date_to) or _in_range(ini.end_date, date_from, date_to)
        ]
    if status:
        initiatives = [ini for ini in initiatives if _status_code(ini) == status]
    return initiatives


@router.get("/reports/initiatives.xlsx")
def report_initiatives_xlsx(
    date_from: Optional[date] = Query(default=None),
    date_to: Optional[date] = Query(default=None),
    status: Optional[str] = Query(default=None),
    department_id: Optional[int] = Query(default=None),
    db: Session = Depends(get_db),
    employee: models.Employee = Depends(require_employee),
):
    initiatives = _filter_initiatives(db, employee, date_from, date_to, status, department_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Инициативы"

    ws.append(REPORT_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for ini in initiatives:
        ws.append(_row_values(ini))

    summary_rows = _summary_rows(initiatives)
    for row in summary_rows:
        ws.append(row)
    # "Итого по отчёту" is the second summary row (right after the blank spacer).
    ws.cell(row=ws.max_row - len(summary_rows) + 2, column=1).font = Font(bold=True)

    widths = [30, 40, 18, 20, 32, 14, 14, 45, 30, 18, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"report_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/reports/initiatives.csv")
def report_initiatives_csv(
    date_from: Optional[date] = Query(default=None),
    date_to: Optional[date] = Query(default=None),
    status: Optional[str] = Query(default=None),
    department_id: Optional[int] = Query(default=None),
    db: Session = Depends(get_db),
    employee: models.Employee = Depends(require_employee),
):
    initiatives = _filter_initiatives(db, employee, date_from, date_to, status, department_id)

    buf = io.StringIO()
    # Excel only auto-detects UTF-8 CSVs (rather than mangling Cyrillic) when
    # the file starts with a BOM and fields are ;-separated in ru-RU locale.
    buf.write("﻿")
    writer = csv.writer(buf, delimiter=";", lineterminator="\r\n")
    writer.writerow(REPORT_HEADERS)
    for ini in initiatives:
        writer.writerow(_row_values(ini))
    for row in _summary_rows(initiatives):
        writer.writerow(row)

    filename = f"report_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.csv"
    return Response(
        content=buf.getvalue().encode("utf-8"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
